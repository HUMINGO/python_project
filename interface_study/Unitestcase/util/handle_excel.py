#coding = utf-8

import openpyxl
import os

#封装操作Excel的方法

class Handle_excel:

    def load_excel(self):

        '''
        获取excel
        '''
        open_excel = openpyxl.load_workbook("/Users/my_home/PycharmProjects/interface_study/case_excel/case_1.xlsx")
        return open_excel

    def get_excel_sheet(self, index=None):

        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_data(self, row, clos):
        '''
        获取某一个单元格的内容
        '''
        data = self.get_excel_sheet().cell(row=row, column=clos).value
        return data

    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_excel_sheet().max_row
        return row

    def get_rows_value(self, row):
        '''
        获取行的内容
        '''
        row_list = []

        for i in self.get_excel_sheet()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, rows, cols, value):
        '''
        将测试结果写到Excel中
        '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(rows, cols, value)
        wb.save("/Users/my_home/PycharmProjects/interface_study/case_excel/case_1.xlsx")

excel_data = Handle_excel()

if __name__ == "__main__":
    handle = Handle_excel()
    # print(handle.get_cell_data(1, 3))
    print(handle.get_rows_value(2))
    handle.excel_write_data(2, 11, "pass")