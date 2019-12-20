#coding = utf-8

'''
操作excel
'''

import openpyxl

wb = openpyxl.load_workbook("/Users/my_home/PycharmProjects/interface_study/case_excel/case_1.xlsx")
all_sheet = wb.sheetnames
# print(all_sheet)

table = wb.get_sheet_by_name('case_01')
print(type(table))

rows = table.max_row   #获取行数
cols = table.max_column  #获取列数
print(rows)
print(cols)
#获取单元格值：
Data = table.cell(row=1, column=2).value
print(Data)